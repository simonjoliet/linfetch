#!/usr/bin/env python3
import time
import openpyxl
import sys, getopt

#Specify I/O files
inputfile = "./sample/in_file.xlsx"
outputfile = "./sample/out_file.xlsx"

#Create a Chrome object
from selenium import webdriver
browser = webdriver.Chrome('chromedriver')

#Open login page
browser.get('https://www.linkedin.com/login?fromSignIn=true&trk=guest_homepage-basic_nav-header-signin')

#Enter login info & submit form
elementID = browser.find_element_by_id('username')
elementID.send_keys('<Your LinkedIn Username>')
elementID = browser.find_element_by_id('password')
elementID.send_keys('<Your LinkedIn Password>')
elementID.submit()

#Open Input file
book = openpyxl.load_workbook(inputfile)
sheet = book.active
max_row = sheet.max_row

#Create Output file
newBook = openpyxl.Workbook()
newSheet = newBook.active
formatSheet = newBook.create_sheet("Format Sheet")
newSheet.name = "User Fetch"

#Writing Headers into the out file
headers = ["First Name", "Last Name", "LinkedIn URL", "LinkedIn Username", "Position", "Company", "LinkedIn Location", "Found in USA"]
 
for col, val in enumerate(headers, start=1):
	newSheet.cell(row=1, column=col).value = val

#Writing Headers into the out file
headers = ["first name", "last name", "domain", "position", "company"]
 
for col, val in enumerate(headers, start=1):
   formatSheet.cell(row=1, column=col).value = val

#set offset
offset = 2

#Iterate each row in the input file
for row in range (offset,max_row +1):
		
	#Instanciate user vars
	i = 0
	hasLink = False
	urlUser = ""
	userName = ""
	position = ""
	companyName = ""
	LinkedInAdress = ""
	FoundInUS = "False"
	try:
		firstName = (sheet.cell(row, 1).value).split()[0].lower().strip()
	except AttributeError:
		firstName = " "
			
	lastName = (sheet.cell(row, 2).value).lower().strip()
	
	#Open the search page
	browser.get("https://www.linkedin.com/search/results/people/?geoUrn=%5B\"103644278\"%5D&keywords=" + firstName + "%20" + lastName + "&origin=FACETED_SEARCH")
	time.sleep(0.6)
	elemsLink = browser.find_elements_by_xpath("//a[@href]")

	#Check if at least one link matches a profile, and if the first letter match the first name
	for elemTmp in elemsLink:
		href = elemTmp.get_attribute("href")
		if href.startswith("https://www.linkedin.com/in/" + firstName[0]):
			hasLink = True
			FoundInUS = "True"

	#If no link was found, the user was probably not in the US. Search globaly
	if hasLink == False:
		browser.get("https://www.linkedin.com/search/results/people/?keywords=" + firstName + "%20" + lastName + "&origin=FACETED_SEARCH")
		time.sleep(0.5)
		elemsLink = browser.find_elements_by_xpath("//a[@href]")

	#Get the Position and Adress HTML element in an array
	elemsPosition = browser.find_elements_by_xpath('//div[@class="entity-result__primary-subtitle t-14 t-black"]')
	elemsAdress = browser.find_elements_by_xpath('//div[@class="entity-result__secondary-subtitle t-14"]')
	
	#Go through all the links found
	for elemLink in elemsLink:
		
		#When the matching link is found
		href = elemLink.get_attribute("href")
		if href.startswith("https://www.linkedin.com/in/" + firstName[0]):
			
			#The url of the user is the link
			urlUser = href
			userName = urlUser.split("/", 4)[-1]
			
			#getting the Position and the Adress the the i-th position as well
			try:
				position = elemsPosition[i].get_attribute('innerHTML').strip().replace("<!---->","")
			except IndexError:
				position = " "

			try:
				LinkedInAdress = elemsAdress[i].get_attribute('innerHTML').strip().replace("<!---->","")
			except IndexError:
				LinkedInAdress = ""

			i += 1

			#Split the name of the company and position if there is an " at " in position
			if " at " in position:
				companyName = position.split(" at ",1)[1] 
				position = position.split(" at ",1)[0] 

			break

	#Writing current row into the out file
	myRow = [sheet.cell(row, 1).value, sheet.cell(row, 2).value, urlUser, userName, position, companyName, LinkedInAdress, FoundInUS, ""]

	#Writing current row into the output file
	for col, val in enumerate(myRow, start=1):
		newSheet.cell(row=row, column=col).value = val

	myRow = ['=PROPER(INDIRECT("Sheet!A"&ROW()))', '=PROPER(INDIRECT("Sheet!B"&ROW()))', "", '=PROPER(INDIRECT("Sheet!E"&ROW()))', '=PROPER(INDIRECT("Sheet!F"&ROW()))']

	for col, val in enumerate(myRow, start=1):
		formatSheet.cell(row=row, column=col).value = val

	#Save the current version of the file
	newBook.save(outputfile)

	#Print debug
	print (str(row) + "=>(#" + str(i) + ")=>USA=" +FoundInUS+"=>" + userName )

browser.quit()